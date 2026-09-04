# -*- coding: utf-8 -*-
"""
Excel Structure Extractor
Извлечение структуры Excel-файлов: листы, ячейки, формулы, значения
"""
import os
import logging
from typing import Dict, List, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

logger = logging.getLogger(__name__)

class ExcelStructureExtractor:
    """
    Извлечение структуры Excel-файлов для дальнейшего маппинга
    """
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls', '.xlsm']
    
    def extract(self, file_path: str) -> Dict[str, Any]:
        """
        Извлечение полной структуры Excel-файла
        
        Args:
            file_path: Путь к Excel-файлу
            
        Returns:
            Dict со структурой файла
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Файл не найден: {file_path}")
            
            ext = Path(file_path).suffix.lower()
            if ext not in self.supported_extensions:
                raise ValueError(f"Неподдерживаемый формат: {ext}")
            
            logger.info(f"Извлечение структуры из: {file_path}")
            
            # Загружаем workbook
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            # Извлекаем структуру
            structure = {
                'file_name': os.path.basename(file_path),
                'sheets': [],
                'summary': {
                    'total_sheets': len(workbook.sheetnames),
                    'has_formulas': False,
                    'has_data': False
                }
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_structure = self._extract_sheet_structure(sheet)
                structure['sheets'].append(sheet_structure)
                
                # Обновляем сводку
                if sheet_structure['has_formulas']:
                    structure['summary']['has_formulas'] = True
                if sheet_structure['data_cells'] > 0:
                    structure['summary']['has_data'] = True
            
            logger.info(f"Извлечено {len(structure['sheets'])} листов")
            return structure
            
        except Exception as e:
            logger.error(f"Ошибка при извлечении структуры: {e}")
            raise
    
    def _extract_sheet_structure(self, sheet) -> Dict[str, Any]:
        """
        Извлечение структуры отдельного листа
        """
        sheet_structure = {
            'name': sheet.title,
            'dimensions': {
                'max_row': sheet.max_row,
                'max_column': sheet.max_column
            },
            'cells': [],
            'formulas': [],
            'data_cells': 0,
            'has_formulas': False,
            'has_headers': False
        }
        
        # Определяем заголовки (первая строка)
        headers = []
        for col in range(1, min(sheet.max_column + 1, 20)):  # Ограничим для производительности
            cell = sheet.cell(row=1, column=col)
            if cell.value:
                headers.append(str(cell.value))
                sheet_structure['has_headers'] = True
        
        # Извлекаем ячейки с данными и формулами
        for row in range(1, min(sheet.max_row + 1, 100)):  # Ограничим для производительности
            for col in range(1, min(sheet.max_column + 1, 20)):
                cell = sheet.cell(row=row, column=col)
                
                if cell.value is not None:
                    cell_data = {
                        'row': row,
                        'col': col,
                        'col_letter': get_column_letter(col),
                        'address': f"{get_column_letter(col)}{row}",
                        'value': str(cell.value) if not isinstance(cell.value, (int, float)) else cell.value,
                        'data_type': self._get_cell_type(cell)
                    }
                    
                    # Проверяем наличие формулы
                    if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell_data['formula'] = cell.value
                        sheet_structure['formulas'].append(cell_data)
                        sheet_structure['has_formulas'] = True
                    else:
                        sheet_structure['data_cells'] += 1
                    
                    sheet_structure['cells'].append(cell_data)
        
        # Определяем потенциальные входные/выходные ячейки
        sheet_structure['input_cells'] = self._identify_input_cells(sheet_structure)
        sheet_structure['output_cells'] = self._identify_output_cells(sheet_structure)
        
        return sheet_structure
    
    def _get_cell_type(self, cell) -> str:
        """Определение типа ячейки"""
        if cell.data_type == 'f':
            return 'formula'
        elif isinstance(cell.value, (int, float)):
            return 'number'
        elif isinstance(cell.value, bool):
            return 'boolean'
        elif isinstance(cell.value, str):
            return 'text'
        else:
            return 'unknown'
    
    def _identify_input_cells(self, sheet_structure: Dict) -> List[Dict]:
        """
        Идентификация потенциальных входных ячеек
        (обычно ячейки без формул, но с данными)
        """
        input_cells = []
        for cell in sheet_structure['cells']:
            if cell['data_type'] != 'formula' and cell['data_type'] != 'boolean':
                # Ячейки с текстом или числами могут быть входами
                input_cells.append({
                    'address': cell['address'],
                    'value': cell['value'],
                    'type': cell['data_type'],
                    'row': cell['row'],
                    'col': cell['col']
                })
        return input_cells
    
    def _identify_output_cells(self, sheet_structure: Dict) -> List[Dict]:
        """
        Идентификация потенциальных выходных ячеек
        (обычно ячейки с формулами или числами в конце таблицы)
        """
        output_cells = []
        for cell in sheet_structure['cells']:
            if cell['data_type'] == 'formula':
                output_cells.append({
                    'address': cell['address'],
                    'formula': cell.get('formula', ''),
                    'value': cell['value'],
                    'row': cell['row'],
                    'col': cell['col']
                })
        return output_cells
    
    def extract_as_dataframe(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Извлечение Excel как DataFrame (для pandas)
        """
        try:
            excel_file = pd.ExcelFile(file_path)
            dataframes = {}
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                dataframes[sheet_name] = df
            
            return dataframes
        except Exception as e:
            logger.error(f"Ошибка при извлечении DataFrame: {e}")
            return {}
    
    def get_sheet_preview(self, file_path: str, max_rows: int = 10) -> Dict[str, Any]:
        """
        Получение предпросмотра всех листов
        """
        try:
            structure = self.extract(file_path)
            preview = {
                'file_name': structure['file_name'],
                'sheets': {}
            }
            
            dataframes = self.extract_as_dataframe(file_path)
            for sheet_name, df in dataframes.items():
                preview['sheets'][sheet_name] = {
                    'preview': df.head(max_rows).to_dict('records'),
                    'shape': df.shape,
                    'columns': df.columns.tolist(),
                    'dtypes': df.dtypes.astype(str).to_dict()
                }
            
            return preview
        except Exception as e:
            logger.error(f"Ошибка при получении предпросмотра: {e}")
            return {}
