# -*- coding: utf-8 -*-
import subprocess
import os
import tempfile
import openpyxl
import logging
from typing import Dict, Any, Optional

logger = logging.getLogger(__name__)

class LibreOfficeCalculator:
    def __init__(self):
        self.is_available = self._check_availability()
    
    def _check_availability(self) -> bool:
        try:
            result = subprocess.run(['libreoffice', '--version'], capture_output=True, timeout=5)
            return result.returncode == 0
        except:
            return False
    
    def calculate_roi(self, file_path: str, mapping: Dict[str, str], business_case: Dict) -> Optional[float]:
        if not self.is_available:
            logger.warning("⚠️ LibreOffice не найден, используем fallback")
            return self._calculate_fallback(file_path, mapping, business_case)
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            sheet = wb.active
            for key, address in mapping.items():
                if key in ['current_costs', 'team_size', 'time_saved', 'hourly_rate']:
                    value = business_case.get(key)
                    if value:
                        sheet[address] = value
            
            temp_path = file_path.replace('.xlsx', '_calc.xlsx')
            wb.save(temp_path)
            
            output_dir = tempfile.mkdtemp()
            result = subprocess.run([
                'libreoffice', '--headless', '--convert-to', 'xlsx',
                '--outdir', output_dir, temp_path
            ], capture_output=True, timeout=30)
            
            if result.returncode != 0:
                logger.error(f"LibreOffice ошибка: {result.stderr}")
                return self._calculate_fallback(file_path, mapping, business_case)
            
            converted_path = os.path.join(output_dir, os.path.basename(temp_path))
            if os.path.exists(converted_path):
                wb_result = openpyxl.load_workbook(converted_path, data_only=True)
                sheet_result = wb_result.active
                roi_value = sheet_result[mapping.get("roi_result", "E5")].value
                os.unlink(temp_path)
                os.unlink(converted_path)
                os.rmdir(output_dir)
                return roi_value
            return None
        except Exception as e:
            logger.error(f"Ошибка LibreOffice: {e}")
            return self._calculate_fallback(file_path, mapping, business_case)
    
    def _calculate_fallback(self, file_path: str, mapping: Dict[str, str], business_case: Dict) -> Optional[float]:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            sheet = wb.active
            for key, address in mapping.items():
                if key in ['current_costs', 'team_size', 'time_saved', 'hourly_rate']:
                    value = business_case.get(key)
                    if value:
                        sheet[address] = value
            temp_path = file_path.replace('.xlsx', '_temp.xlsx')
            wb.save(temp_path)
            try:
                from pycel import ExcelCompiler
                compiler = ExcelCompiler(temp_path)
                roi_value = compiler.evaluate(mapping.get("roi_result", "E5"))
                os.unlink(temp_path)
                return roi_value
            except ImportError:
                wb_calc = openpyxl.load_workbook(temp_path, data_only=True)
                sheet_calc = wb_calc.active
                roi_value = sheet_calc[mapping.get("roi_result", "E5")].value
                os.unlink(temp_path)
                return roi_value
        except Exception as e:
            logger.error(f"Fallback ошибка: {e}")
            return None
