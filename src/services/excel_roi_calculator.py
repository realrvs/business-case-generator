# -*- coding: utf-8 -*-
"""
Excel ROI Calculator
Подстановка данных и пересчет Excel-моделей
"""
import os
import logging
import tempfile
import shutil
import time
from typing import Dict, List, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

logger = logging.getLogger(__name__)

class ExcelROICalculator:
    """
    Подстановка данных в Excel-модель и пересчет формул
    """
    
    def __init__(self):
        self.temp_dir = tempfile.mkdtemp(prefix="excel_roi_")
        logger.info(f"Временная директория создана: {self.temp_dir}")
    
    def calculate(self, file_path: str, data: Dict[str, Any], mapping: Dict[str, Any]) -> Dict[str, Any]:
        """
        Подстановка данных и пересчет Excel-модели
        """
        workbook = None
        temp_file_path = None
        
        try:
            logger.info(f"Начало расчета для файла: {file_path}")
            
            # Копируем файл во временную директорию
            temp_file_path = self._copy_to_temp(file_path)
            
            # Загружаем workbook
            workbook = openpyxl.load_workbook(temp_file_path)
            
            # Подставляем данные
            self._apply_data(workbook, data, mapping)
            
            # Сохраняем изменения
            workbook.save(temp_file_path)
            workbook.close()
            workbook = None
            
            # Пересчитываем формулы (если есть LibreOffice)
            if self._is_libreoffice_available():
                recalculated_path = self._recalculate_with_libreoffice(temp_file_path)
                if recalculated_path:
                    temp_file_path = recalculated_path
            
            # Извлекаем результаты
            results = self._extract_results(temp_file_path, mapping)
            output_values = self._get_output_values(temp_file_path, mapping)
            
            return {
                'status': 'success',
                'results': results,
                'output_cells': output_values,
                'applied_data': data,
                'mapping_used': mapping
            }
            
        except Exception as e:
            logger.error(f"Ошибка при расчете: {e}")
            return {
                'status': 'error',
                'error': str(e)
            }
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except:
                    pass
    
    def _copy_to_temp(self, file_path: str) -> str:
        """Копирование файла во временную директорию"""
        file_name = os.path.basename(file_path)
        temp_path = os.path.join(self.temp_dir, f"calc_{file_name}")
        shutil.copy2(file_path, temp_path)
        return temp_path
    
    def _apply_data(self, workbook, data: Dict[str, Any], mapping: Dict[str, Any]):
        """Подстановка данных в ячейки"""
        mapped_cells = mapping.get('mapped_cells', {})
        
        for key, cell_info in mapped_cells.items():
            if cell_info.get('type') == 'input':
                try:
                    sheet_name, address = key.split('!')
                    sheet = workbook[sheet_name]
                    cell = sheet[address]
                    
                    cell_name = cell_info.get('name', '')
                    value = data.get(cell_name)
                    
                    if value is not None:
                        cell.value = value
                        logger.info(f"Подставлено значение {value} в {key}")
                        
                except Exception as e:
                    logger.warning(f"Не удалось подставить данные в {key}: {e}")
    
    def _is_libreoffice_available(self) -> bool:
        """Проверка доступности LibreOffice"""
        import shutil
        return shutil.which('soffice') is not None
    
    def _recalculate_with_libreoffice(self, file_path: str) -> Optional[str]:
        """Пересчет формул через LibreOffice"""
        try:
            import subprocess
            
            output_dir = os.path.dirname(file_path)
            
            cmd = [
                'soffice',
                '--headless',
                '--convert-to', 'xlsx',
                '--outdir', output_dir,
                file_path
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            
            if result.returncode == 0:
                for f in os.listdir(output_dir):
                    if f.endswith('.xlsx') and f != os.path.basename(file_path):
                        return os.path.join(output_dir, f)
            
            logger.warning(f"LibreOffice пересчет не удался: {result.stderr}")
            return None
            
        except Exception as e:
            logger.error(f"Ошибка при вызове LibreOffice: {e}")
            return None
    
    def _extract_results(self, file_path: str, mapping: Dict) -> Dict[str, Any]:
        """Извлечение результатов из модели"""
        workbook = None
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            results = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in range(1, min(sheet.max_row + 1, 50)):
                    row_data = []
                    for col in range(1, min(sheet.max_column + 1, 20)):
                        cell = sheet.cell(row=row, column=col)
                        row_data.append({
                            'address': f"{get_column_letter(col)}{row}",
                            'value': cell.value
                        })
                    sheet_data.append(row_data)
                
                results[sheet_name] = sheet_data
            
            return results
            
        except Exception as e:
            logger.error(f"Ошибка при извлечении результатов: {e}")
            return {}
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except:
                    pass
    
    def _get_output_values(self, file_path: str, mapping: Dict) -> Dict[str, Any]:
        """Получение значений выходных ячеек"""
        workbook = None
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            output_values = {}
            
            mapped_cells = mapping.get('mapped_cells', {})
            for key, cell_info in mapped_cells.items():
                if cell_info.get('type') == 'output':
                    try:
                        sheet_name, address = key.split('!')
                        sheet = workbook[sheet_name]
                        cell = sheet[address]
                        output_values[key] = {
                            'value': cell.value,
                            'name': cell_info.get('name', ''),
                            'description': cell_info.get('description', '')
                        }
                    except Exception as e:
                        logger.warning(f"Не удалось прочитать {key}: {e}")
            
            return output_values
            
        except Exception as e:
            logger.error(f"Ошибка при чтении выходных ячеек: {e}")
            return {}
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except:
                    pass
    
    def cleanup(self):
        """
        Очистка временной директории с повторными попытками
        """
        if not os.path.exists(self.temp_dir):
            return
            
        max_attempts = 5
        for attempt in range(max_attempts):
            try:
                # Увеличиваем задержку с каждой попыткой
                time.sleep(0.3 * (attempt + 1))
                shutil.rmtree(self.temp_dir, ignore_errors=True)
                
                # Проверяем, удалилось ли
                if not os.path.exists(self.temp_dir):
                    logger.info(f"Временная директория удалена: {self.temp_dir}")
                    return
                    
            except Exception as e:
                logger.warning(f"Попытка {attempt + 1} удаления не удалась: {e}")
        
        # Если после всех попыток не удалилось, пробуем принудительно
        try:
            if os.path.exists(self.temp_dir):
                # Удаляем все файлы внутри вручную
                for root, dirs, files in os.walk(self.temp_dir, topdown=False):
                    for name in files:
                        try:
                            os.remove(os.path.join(root, name))
                        except:
                            pass
                    for name in dirs:
                        try:
                            os.rmdir(os.path.join(root, name))
                        except:
                            pass
                os.rmdir(self.temp_dir)
                logger.info(f"Временная директория принудительно удалена: {self.temp_dir}")
        except Exception as e:
            logger.error(f"Не удалось удалить временную директорию: {e}")
